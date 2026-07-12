# -*- coding: utf-8 -*-
"""蓄電池設置 予実進捗管理シート (.xlsx) を生成するスクリプト。
openpyxl のみ使用。数式・条件付き書式・入力規則入りで、そのまま実務に使える構成。
再生成したい場合:  python3 build_xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

OUT = "蓄電池設置_予実進捗管理シート.xlsx"

PHASES = ["受注・契約", "現地調査", "設計・図面", "補助金申請", "系統連系申請",
          "機器発注・納品", "設置工事", "完了検査・連系", "引渡し"]
NP = len(PHASES)                     # 9
FIRST_DATA = 3                       # 明細開始行
NROWS = 40                           # 明細行数
LAST_DATA = FIRST_DATA + NROWS - 1

# 列レイアウト
# A:No B:案件名 C:顧客 D:担当 E:容量 F:受注日
# G..O 予定日(9)  P..X 実績日(9)  Y:進捗率 Z:遅延 AA:ステータス AB:次期限 AC:メモ
COL_NO, COL_NAME, COL_CUST, COL_OWN, COL_CAP, COL_ORDER = 1, 2, 3, 4, 5, 6
PLAN0 = 7                            # G
ACT0 = PLAN0 + NP                    # P
COL_PCT = ACT0 + NP                  # Y
COL_DELAY = COL_PCT + 1             # Z
COL_STATUS = COL_DELAY + 1          # AA
COL_NEXT = COL_STATUS + 1           # AB
COL_MEMO = COL_NEXT + 1             # AC
LAST_COL = COL_MEMO

def L(c):  # 列番号→英字
    return get_column_letter(c)

# ---- 色・スタイル ----
INK = "1F2937"
NAVY = "1E293B"
BLUE = "2563EB"
GREY_HDR = "F1F5F9"
fill_navy = PatternFill("solid", fgColor=NAVY)
fill_plan = PatternFill("solid", fgColor="EAF1FB")   # 予定ブロック薄青
fill_act = PatternFill("solid", fgColor="EAF7EE")    # 実績ブロック薄緑
fill_hdr = PatternFill("solid", fgColor=GREY_HDR)
fill_alt = PatternFill("solid", fgColor="FAFBFD")

thin = Side(style="thin", color="D9DEE8")
med = Side(style="medium", color="B7C0D0")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

f_title = Font(name="Meiryo", size=15, bold=True, color="FFFFFF")
f_sub = Font(name="Meiryo", size=9, color="C7D2E0")
f_hdr = Font(name="Meiryo", size=9, bold=True, color=INK)
f_grp = Font(name="Meiryo", size=10, bold=True, color="FFFFFF")
f_cell = Font(name="Meiryo", size=10, color=INK)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center")
rightA = Alignment(horizontal="right", vertical="center")

wb = Workbook()

# =====================================================================
#  シート1: 進捗管理
# =====================================================================
ws = wb.active
ws.title = "進捗管理"
ws.sheet_view.showGridLines = False

# --- タイトル帯 (行1) ---
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_ORDER)
c = ws.cell(1, 1, "🔋 蓄電池設置 予実進捗管理シート")
c.font = f_title; c.fill = fill_navy; c.alignment = Alignment(horizontal="left", vertical="center")
# 予定/実績 グループ見出し (行1)
ws.merge_cells(start_row=1, start_column=PLAN0, end_row=1, end_column=PLAN0+NP-1)
g = ws.cell(1, PLAN0, "予 定 日"); g.font = f_grp; g.alignment = center
g.fill = PatternFill("solid", fgColor=BLUE)
ws.merge_cells(start_row=1, start_column=ACT0, end_row=1, end_column=ACT0+NP-1)
g = ws.cell(1, ACT0, "実 績 日"); g.font = f_grp; g.alignment = center
g.fill = PatternFill("solid", fgColor="16A34A")
ws.merge_cells(start_row=1, start_column=COL_PCT, end_row=1, end_column=COL_MEMO)
g = ws.cell(1, COL_PCT, "自動計算（数式）"); g.font = f_sub; g.alignment = center
g.fill = fill_navy
_colored = set(range(1, COL_ORDER+1)) | set(range(PLAN0, PLAN0+NP)) | set(range(ACT0, ACT0+NP)) | set(range(COL_PCT, COL_MEMO+1))
for col in range(1, LAST_COL+1):
    if col not in _colored:
        ws.cell(1, col).fill = fill_navy
ws.row_dimensions[1].height = 26

# --- 見出し (行2) ---
headers = {COL_NO:"No", COL_NAME:"案件名", COL_CUST:"顧客名", COL_OWN:"担当",
           COL_CAP:"容量\n(kWh)", COL_ORDER:"受注日",
           COL_PCT:"進捗率", COL_DELAY:"遅延\n工程", COL_STATUS:"ステータス",
           COL_NEXT:"次の期限", COL_MEMO:"メモ / 備考"}
for i, name in enumerate(PHASES):
    headers[PLAN0+i] = f"{i+1}. {name}"
    headers[ACT0+i] = f"{i+1}. {name}"
for col, txt in headers.items():
    cell = ws.cell(2, col, txt)
    cell.font = f_hdr; cell.alignment = center; cell.border = border_all
    if PLAN0 <= col < PLAN0+NP:
        cell.fill = fill_plan
    elif ACT0 <= col < ACT0+NP:
        cell.fill = fill_act
    else:
        cell.fill = fill_hdr
ws.row_dimensions[2].height = 40

# --- 明細行の数式 ---
date_style = "yyyy/m/d"
for r in range(FIRST_DATA, LAST_DATA+1):
    plan_rng = f"{L(PLAN0)}{r}:{L(PLAN0+NP-1)}{r}"
    act_rng = f"{L(ACT0)}{r}:{L(ACT0+NP-1)}{r}"
    name_ref = f"$B{r}"

    # 入力セルの体裁
    for col in range(1, LAST_COL+1):
        cell = ws.cell(r, col)
        cell.border = border_all
        cell.font = f_cell
        if col == COL_NO:
            cell.value = r - FIRST_DATA + 1
            cell.alignment = center
        elif col in (COL_ORDER,) or (PLAN0 <= col < PLAN0+NP) or (ACT0 <= col < ACT0+NP):
            cell.number_format = date_style
            cell.alignment = center
        elif col == COL_CAP:
            cell.alignment = center
        else:
            cell.alignment = left
        if PLAN0 <= col < PLAN0+NP:
            cell.fill = fill_plan
        elif ACT0 <= col < ACT0+NP:
            cell.fill = fill_act

    # 進捗率 = 実績入力数 / 工程数
    ws.cell(r, COL_PCT).value = f'=IF({name_ref}="","",COUNT({act_rng})/{NP})'
    ws.cell(r, COL_PCT).number_format = "0%"
    ws.cell(r, COL_PCT).alignment = center
    # 遅延工程数 = 予定を過ぎて未実績の工程
    ws.cell(r, COL_DELAY).value = (
        f'=IF({name_ref}="","",SUMPRODUCT(({act_rng}="")*({plan_rng}>0)*({plan_rng}<TODAY())))'
    )
    ws.cell(r, COL_DELAY).alignment = center
    # ステータス
    ws.cell(r, COL_STATUS).value = (
        f'=IF({name_ref}="","",'
        f'IF({L(COL_PCT)}{r}=1,"完了",'
        f'IF({L(COL_DELAY)}{r}>0,"遅延",'
        f'IF(COUNT({plan_rng},{act_rng})>0,"進行中","未着手"))))'
    )
    ws.cell(r, COL_STATUS).alignment = center
    ws.cell(r, COL_STATUS).font = Font(name="Meiryo", size=10, bold=True, color=INK)
    # 次の期限 = 未実績の工程のうち最も早い予定日 (AGGREGATE: 非配列・広い互換性)
    ws.cell(r, COL_NEXT).value = (
        f'=IF(OR({name_ref}="",SUMPRODUCT(({act_rng}="")*({plan_rng}>0))=0),"",'
        f'AGGREGATE(15,6,{plan_rng}/(({act_rng}="")*({plan_rng}>0)),1))'
    )
    ws.cell(r, COL_NEXT).number_format = date_style
    ws.cell(r, COL_NEXT).alignment = center

    ws.row_dimensions[r].height = 20

# --- サンプルデータ (2件) ---
today = date(2026, 7, 12)
def d(off): return today + timedelta(days=off)
samples = [
    dict(name="サンプル：〇〇邸 蓄電池設置工事", cust="山田 太郎", own="鈴木", cap=9.8,
         order=d(-30), plans=[d(-25+i*5) for i in range(NP)],
         acts=[d(-24+i*5) for i in range(4)] + [None]*5,
         memo="サンプルデータ。不要なら行を削除してください"),
    dict(name="サンプル：△△マンション 共用部", cust="□□管理組合", own="佐藤", cap=16.4,
         order=d(-10), plans=[d(-8+i*6) for i in range(NP)],
         acts=[d(-7), d(-1)] + [None]*7,
         memo="補助金：DR補助金 申請予定"),
]
for si, s in enumerate(samples):
    r = FIRST_DATA + si
    ws.cell(r, COL_NAME, s["name"])
    ws.cell(r, COL_CUST, s["cust"])
    ws.cell(r, COL_OWN, s["own"])
    ws.cell(r, COL_CAP, s["cap"])
    ws.cell(r, COL_ORDER, s["order"])
    ws.cell(r, COL_MEMO, s["memo"])
    for i in range(NP):
        if s["plans"][i]:
            ws.cell(r, PLAN0+i, s["plans"][i])
        if s["acts"][i]:
            ws.cell(r, ACT0+i, s["acts"][i])

# --- 列幅 ---
ws.column_dimensions[L(COL_NO)].width = 4.5
ws.column_dimensions[L(COL_NAME)].width = 26
ws.column_dimensions[L(COL_CUST)].width = 12
ws.column_dimensions[L(COL_OWN)].width = 7
ws.column_dimensions[L(COL_CAP)].width = 7
ws.column_dimensions[L(COL_ORDER)].width = 11
for i in range(NP):
    ws.column_dimensions[L(PLAN0+i)].width = 10
    ws.column_dimensions[L(ACT0+i)].width = 10
ws.column_dimensions[L(COL_PCT)].width = 8
ws.column_dimensions[L(COL_DELAY)].width = 6
ws.column_dimensions[L(COL_STATUS)].width = 9
ws.column_dimensions[L(COL_NEXT)].width = 11
ws.column_dimensions[L(COL_MEMO)].width = 30

# --- フリーズ & オートフィルタ ---
ws.freeze_panes = f"{L(COL_ORDER+1)}{FIRST_DATA}"     # 受注日まで固定、明細から下スクロール
ws.auto_filter.ref = f"A2:{L(LAST_COL)}{LAST_DATA}"

# --- 条件付き書式 ---
data_rng_plan = f"{L(PLAN0)}{FIRST_DATA}:{L(PLAN0+NP-1)}{LAST_DATA}"
data_rng_act = f"{L(ACT0)}{FIRST_DATA}:{L(ACT0+NP-1)}{LAST_DATA}"
status_rng = f"{L(COL_STATUS)}{FIRST_DATA}:{L(COL_STATUS)}{LAST_DATA}"
pct_rng = f"{L(COL_PCT)}{FIRST_DATA}:{L(COL_PCT)}{LAST_DATA}"
delay_rng = f"{L(COL_DELAY)}{FIRST_DATA}:{L(COL_DELAY)}{LAST_DATA}"

red_fill = PatternFill("solid", fgColor="FDE2E2"); red_font = Font(color="B91C1C", bold=True, name="Meiryo", size=10)
grn_fill = PatternFill("solid", fgColor="DCFCE7"); grn_font = Font(color="15803D", bold=True, name="Meiryo", size=10)
blu_fill = PatternFill("solid", fgColor="E0EBFE"); blu_font = Font(color="1D4ED8", bold=True, name="Meiryo", size=10)
gry_fill = PatternFill("solid", fgColor="EEF1F6"); gry_font = Font(color="64748B", bold=True, name="Meiryo", size=10)
org_fill = PatternFill("solid", fgColor="FEF0D9")

# ステータス色分け
ws.conditional_formatting.add(status_rng, FormulaRule(formula=[f'{L(COL_STATUS)}{FIRST_DATA}="完了"'], fill=grn_fill, font=grn_font))
ws.conditional_formatting.add(status_rng, FormulaRule(formula=[f'{L(COL_STATUS)}{FIRST_DATA}="遅延"'], fill=red_fill, font=red_font))
ws.conditional_formatting.add(status_rng, FormulaRule(formula=[f'{L(COL_STATUS)}{FIRST_DATA}="進行中"'], fill=blu_fill, font=blu_font))
ws.conditional_formatting.add(status_rng, FormulaRule(formula=[f'{L(COL_STATUS)}{FIRST_DATA}="未着手"'], fill=gry_fill, font=gry_font))
# 遅延工程数 > 0 → 赤
ws.conditional_formatting.add(delay_rng, FormulaRule(formula=[f'{L(COL_DELAY)}{FIRST_DATA}>0'], fill=red_fill, font=red_font))
# 予定日を過ぎて未実績 → 予定セル赤 (対の実績セルは +NP 列右)
ws.conditional_formatting.add(
    data_rng_plan,
    FormulaRule(formula=[f'AND({L(PLAN0)}{FIRST_DATA}<>"",{L(PLAN0)}{FIRST_DATA}<TODAY(),{L(ACT0)}{FIRST_DATA}="")'],
                fill=red_fill))
# 実績が予定より遅れた → 実績セル橙 (対の予定セルは -NP 列左)
ws.conditional_formatting.add(
    data_rng_act,
    FormulaRule(formula=[f'AND({L(ACT0)}{FIRST_DATA}<>"",{L(PLAN0)}{FIRST_DATA}<>"",{L(ACT0)}{FIRST_DATA}>{L(PLAN0)}{FIRST_DATA})'],
                fill=org_fill))
# 進捗率データバー
ws.conditional_formatting.add(pct_rng, DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                                                   color="22C55E", showValue=True))

# =====================================================================
#  シート2: ダッシュボード
# =====================================================================
db = wb.create_sheet("ダッシュボード")
db.sheet_view.showGridLines = False
db.column_dimensions["A"].width = 3
for col in "BCDEF":
    db.column_dimensions[col].width = 18

db.merge_cells("B2:F2")
t = db.cell(2, 2, "🔋 蓄電池設置 進捗ダッシュボード")
t.font = Font(name="Meiryo", size=14, bold=True, color=INK)

status_full = f"進捗管理!${L(COL_STATUS)}${FIRST_DATA}:${L(COL_STATUS)}${LAST_DATA}"
name_full = f"進捗管理!${L(COL_NAME)}${FIRST_DATA}:${L(COL_NAME)}${LAST_DATA}"
pct_full = f"進捗管理!${L(COL_PCT)}${FIRST_DATA}:${L(COL_PCT)}${LAST_DATA}"

kpis = [
    ("総案件数", f'=COUNTA({name_full})', "2563EB"),
    ("進行中",   f'=COUNTIF({status_full},"進行中")', "1D4ED8"),
    ("遅延あり", f'=COUNTIF({status_full},"遅延")', "B91C1C"),
    ("完了",     f'=COUNTIF({status_full},"完了")', "15803D"),
    ("未着手",   f'=COUNTIF({status_full},"未着手")', "64748B"),
]
r0 = 4
for i, (label, formula, color) in enumerate(kpis):
    col = 2 + i
    cell_l = db.cell(r0, col, label)
    cell_l.font = Font(name="Meiryo", size=10, color="6B7688")
    cell_l.alignment = center
    cell_l.fill = fill_hdr
    cell_l.border = border_all
    cell_n = db.cell(r0+1, col, formula)
    cell_n.font = Font(name="Meiryo", size=22, bold=True, color=color)
    cell_n.alignment = center
    cell_n.border = border_all
db.row_dimensions[r0+1].height = 40

# 平均進捗率
db.cell(r0+3, 2, "平均進捗率")
db.cell(r0+3, 2).font = Font(name="Meiryo", size=10, color="6B7688")
avg = db.cell(r0+3, 3, f'=IFERROR(AVERAGE({pct_full}),0)')
avg.number_format = "0%"; avg.font = Font(name="Meiryo", size=14, bold=True, color=BLUE)

# 使い方メモ
notes = [
    "■ 使い方",
    "1. 「進捗管理」シートに案件を1行ずつ入力します（案件名・顧客・担当・容量・受注日）。",
    "2. 各工程の『予定日』と、完了したら『実績日』を入力します。",
    "3. 進捗率・遅延工程・ステータス・次の期限は自動計算されます（入力不要）。",
    "",
    "■ 色の意味",
    "・予定日セルが赤 … 予定日を過ぎたのに実績日が未入力（遅延）",
    "・実績日セルが橙 … 実績日が予定日より遅れて完了",
    "・ステータス緑=完了 / 赤=遅延 / 青=進行中 / 灰=未着手",
    "",
    "■ 標準工程（9ステップ）",
    "受注・契約 → 現地調査 → 設計・図面 → 補助金申請 → 系統連系申請 →",
    "機器発注・納品 → 設置工事 → 完了検査・連系 → 引渡し",
    "",
    "※ 工程を増減したい場合は「進捗管理」シートの列を追加・削除し、数式範囲を調整してください。",
]
rn = r0 + 6
for i, line in enumerate(notes):
    cell = db.cell(rn+i, 2, line)
    if line.startswith("■"):
        cell.font = Font(name="Meiryo", size=11, bold=True, color=INK)
    else:
        cell.font = Font(name="Meiryo", size=10, color="374151")

wb.save(OUT)
print("wrote", OUT)
